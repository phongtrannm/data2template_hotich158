import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the files
file_goc_path = 'DataKhaiSinh.xlsx'  # Path to the original file
file_template_path = 'TeamplateDangKyKhaiSinh.xlsx'  # Path to the template file

# Read the data
df_goc = pd.read_excel(file_goc_path)
df_template = pd.read_excel(file_template_path)

# Normalize column names by trimming and converting to uppercase for the source file
df_goc.columns = df_goc.columns.str.strip().str.upper()

# Clean and normalize template columns
def clean_column_name(column_name):
    # Remove patterns like "i=[1,2,3,4,5,6]" or "i=[1,2,3,4,5]"
    cleaned_name = re.sub(r'i=\[\d+(,\d+)*\]', '', column_name)
    # Trim and convert to uppercase
    return cleaned_name.strip().upper()

df_template.columns = [clean_column_name(col) for col in df_template.columns]

# Identify columns in data that are not in template
extra_columns = set(df_goc.columns) - set(df_template.columns)
print("Columns present in data but missing in template:", extra_columns)

# Map for special cases
column_mapping = {'NOIDANGKY': 'NOIDANGKY UBND ...'}

# Create a new DataFrame to match the template structure
merged_df = pd.DataFrame()

# Copy data from file_goc to merged_df according to columns in file_template
for column in df_template.columns:
    # Find the corresponding source column, handling special cases
    source_column = next((key for key, value in column_mapping.items() if value == column), column)
    if source_column in df_goc.columns:
        merged_df[column] = df_goc[source_column]
    else:
        # If column does not exist in file_goc, leave it empty
        merged_df[column] = ""

# Add 'STT' column with sequential numbering starting from 1
if 'STT' in merged_df.columns:
    merged_df['STT'] = range(1, len(merged_df) + 1)

# Format 'NGAYDANGKY' column if it exists and has data
if 'NGAYDANGKY' in merged_df.columns:
    merged_df['NGAYDANGKY'] = pd.to_datetime(merged_df['NGAYDANGKY'], errors='coerce', dayfirst=True).dt.strftime('%d.%m.%Y')
    merged_df['NGAYDANGKY'] = merged_df['NGAYDANGKY'].fillna('')  # Ensure empty strings for missing data

# Format 'NKSNGAYSINH' column if it exists and has data
if 'NKSNGAYSINH' in merged_df.columns:
    merged_df['NKSNGAYSINH'] = pd.to_datetime(
        merged_df['NKSNGAYSINH'], errors='coerce', dayfirst=True
    ).dt.strftime('%d.%m.%Y')
    merged_df['NKSNGAYSINH'] = merged_df['NKSNGAYSINH'].fillna('')  # Ensure empty strings for missing data

# Format columns that contain 'SODINHDANH' as text if they contain data
for column in merged_df.columns:
    if re.search(r'SODINHDANH', column):  # Check if 'SODINHDANH' is anywhere in the column name
        # Convert to text format only if the column has non-empty values
        merged_df[column] = merged_df[column].apply(lambda x: str(int(x)) if isinstance(x, float) and x.is_integer() else str(x) if pd.notnull(x) else x)

# Save the merged data to a temporary Excel file to allow formatting
output_path = 'Merged_KhaiSinh_template.xlsx'
merged_df.to_excel(output_path, index=False)

# Open the Excel file to apply styles and adjust column widths
wb = load_workbook(output_path)
ws = wb.active
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Apply red fill to header cells of columns that were missing in the source data
for col_idx, column in enumerate(df_template.columns, start=1):
    source_column = next((key for key, value in column_mapping.items() if value == column), column)
    if source_column not in df_goc.columns:
        ws.cell(row=1, column=col_idx).fill = red_fill

# Adjust column widths based on the maximum width of the data in each column
for col in ws.columns:
    max_length = max(len(str(cell.value)) for cell in col if cell.value)  # Maximum length of data in the column
    adjusted_width = max(max_length + 2, 10)  # Add some padding and set a minimum width
    ws.column_dimensions[col[0].column_letter].width = adjusted_width

# Save the final output
wb.save(output_path)

print(f"Merged data saved to {output_path} with missing columns highlighted in red and adjusted column widths.")
